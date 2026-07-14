#!/usr/bin/env python3
"""
Fixture generator for the MarkItDown deep review.

Produces, deterministically (fixed seed), the fixtures that are not fetched from
the web:
  1. A complex-table HTML matrix: 13 individually-addressable <table> cases that
     stress colspan/rowspan/nested/wide/headerless/malformed/RTL/etc. Each case
     is emitted BOTH as one combined page (complex_tables.html) and as a
     manifest recording, per case, the expected logical shape so the scorer has a
     pre-registered ground truth (anti-bias Part 3: expected set written before
     the run).
  2. A large XLSX for the scale test (large_sheet.xlsx): 50 sheets is overkill;
     we use 1 sheet x 50k rows x 8 cols of deterministic data.
  3. A wide XLSX (wide_sheet.xlsx): 1 sheet x 200 rows x 64 cols.

Everything here is synthetic, no copyright, safe to open-source.
"""
import json
import os
import random

HERE = os.path.dirname(os.path.abspath(__file__))
# Prefer the research-pack layout if present; else the clean public-repo layout.
_pack = os.path.abspath(os.path.join(HERE, "..", "artifacts", "fixtures"))
_repo = os.path.abspath(os.path.join(HERE, "..", "fixtures"))
FIX = _pack if os.path.isdir(_pack) else _repo
TAB = os.path.join(FIX, "tables")
os.makedirs(os.path.join(FIX, "docs"), exist_ok=True)
os.makedirs(TAB, exist_ok=True)

random.seed(1706)  # deterministic

# ----------------------------------------------------------------------------
# 1. Complex-table matrix.  Each entry: (id, human_description, html, expected)
#    expected: a dict of pre-registered structural facts the converter's Markdown
#    is scored against.  "logical_rows"/"logical_cols" describe the human-visible
#    grid AFTER accounting for spans.
# ----------------------------------------------------------------------------
CASES = []


def case(cid, desc, html, expected):
    CASES.append({"id": cid, "desc": desc, "html": html, "expected": expected})


# T01 — plain baseline (control)
case(
    "t01_plain",
    "Plain 3x3 table with a header row (control case).",
    """<table><thead><tr><th>Name</th><th>Role</th><th>City</th></tr></thead>
<tbody>
<tr><td>Ada</td><td>Engineer</td><td>London</td></tr>
<tr><td>Bo</td><td>Designer</td><td>Berlin</td></tr>
</tbody></table>""",
    {"logical_rows": 3, "logical_cols": 3,
     "must_contain": ["Ada", "Engineer", "London", "Bo", "Designer", "Berlin"],
     "note": "baseline; a correct GFM table has header + separator + 2 data rows"},
)

# T02 — colspan in header
case(
    "t02_colspan_header",
    "Header cell spanning 2 columns (colspan=2) over a 3-column body.",
    """<table>
<tr><th colspan="2">Person</th><th rowspan="1">City</th></tr>
<tr><th>First</th><th>Last</th><th></th></tr>
<tr><td>Ada</td><td>Lovelace</td><td>London</td></tr>
<tr><td>Alan</td><td>Turing</td><td>Manchester</td></tr>
</table>""",
    {"logical_rows": 4, "logical_cols": 3,
     "must_contain": ["Person", "First", "Last", "City", "Ada", "Lovelace",
                      "London", "Turing", "Manchester"],
     "spanning": "colspan=2 on 'Person'",
     "note": "spanned header cell — does converter duplicate/blank the spanned column?"},
)

# T03 — rowspan in first column
case(
    "t03_rowspan_col",
    "First column uses rowspan=2 to group two rows under one label.",
    """<table><thead><tr><th>Group</th><th>Item</th><th>Qty</th></tr></thead>
<tbody>
<tr><td rowspan="2">Fruit</td><td>Apple</td><td>5</td></tr>
<tr><td>Banana</td><td>8</td></tr>
<tr><td rowspan="2">Veg</td><td>Carrot</td><td>3</td></tr>
<tr><td>Pea</td><td>9</td></tr>
</tbody></table>""",
    {"logical_rows": 5, "logical_cols": 3,
     "must_contain": ["Group", "Fruit", "Apple", "Banana", "Veg", "Carrot", "Pea"],
     "spanning": "rowspan=2 on 'Fruit' and 'Veg'",
     "note": "rowspan grouping — Markdown has no rowspan; does the label repeat, blank, or drop the second row?"},
)

# T04 — both colspan and rowspan (merged corner)
case(
    "t04_colrowspan",
    "Mixed colspan and rowspan (a merged top-left 2x2 block).",
    """<table>
<tr><td rowspan="2" colspan="2">Merged 2x2</td><td>C</td><td>D</td></tr>
<tr><td>c2</td><td>d2</td></tr>
<tr><td>a3</td><td>b3</td><td>c3</td><td>d3</td></tr>
</table>""",
    {"logical_rows": 3, "logical_cols": 4,
     "must_contain": ["Merged 2x2", "c2", "d2", "a3", "b3", "c3", "d3"],
     "spanning": "rowspan=2 colspan=2 merged block",
     "note": "2D merge — the hardest span case for a flat pipe grid"},
)

# T05 — nested table inside a cell
case(
    "t05_nested",
    "A table nested inside one cell of an outer table.",
    """<table><thead><tr><th>Section</th><th>Detail</th></tr></thead>
<tbody>
<tr><td>Specs</td><td>
  <table><tr><th>Key</th><th>Val</th></tr>
  <tr><td>CPU</td><td>8 core</td></tr>
  <tr><td>RAM</td><td>16 GB</td></tr></table>
</td></tr>
<tr><td>Price</td><td>999</td></tr>
</tbody></table>""",
    {"logical_rows": "outer 3 / inner 3", "logical_cols": "outer 2 / inner 2",
     "must_contain": ["Section", "Detail", "Specs", "CPU", "8 core", "RAM",
                      "16 GB", "Price", "999"],
     "spanning": "nested <table> inside a <td>",
     "note": "GFM cannot nest tables — does inner data survive as text, flatten, or vanish?"},
)

# T06 — very wide table (24 columns)
def _wide24():
    head = "".join(f"<th>C{i}</th>" for i in range(24))
    row1 = "".join(f"<td>r1c{i}</td>" for i in range(24))
    row2 = "".join(f"<td>r2c{i}</td>" for i in range(24))
    return f"<table><tr>{head}</tr><tr>{row1}</tr><tr>{row2}</tr></table>"
case(
    "t06_wide24",
    "Very wide table: 24 columns x 2 data rows.",
    _wide24(),
    {"logical_rows": 3, "logical_cols": 24,
     "must_contain": ["C0", "C23", "r1c0", "r1c23", "r2c0", "r2c23"],
     "note": "width stress — are all 24 columns kept, and is the header/separator arity correct?"},
)

# T07 — headerless table (no <th>, all <td>)
case(
    "t07_headerless",
    "Table with NO header cells — every cell is a <td>.",
    """<table>
<tr><td>2019</td><td>120</td><td>up</td></tr>
<tr><td>2020</td><td>98</td><td>down</td></tr>
<tr><td>2021</td><td>141</td><td>up</td></tr>
</table>""",
    {"logical_rows": 3, "logical_cols": 3,
     "must_contain": ["2019", "120", "2020", "98", "2021", "141"],
     "note": "GFM REQUIRES a header row. Does the converter promote row 1 to header (losing it as data) or synthesize a blank header?"},
)

# T08 — malformed: ragged rows (unequal cell counts)
case(
    "t08_ragged",
    "Malformed table: rows have different numbers of cells (2, 4, 3).",
    """<table>
<tr><td>a</td><td>b</td></tr>
<tr><td>c</td><td>d</td><td>e</td><td>f</td></tr>
<tr><td>g</td><td>h</td><td>i</td></tr>
</table>""",
    {"logical_rows": 3, "logical_cols": "ragged (2/4/3)",
     "must_contain": ["a", "b", "c", "d", "e", "f", "g", "h", "i"],
     "note": "ragged rows — does the converter pad, truncate to the first row's width, or emit invalid GFM?"},
)

# T09 — empty cells and whitespace cells
case(
    "t09_empty_cells",
    "Table with truly empty cells and whitespace-only cells.",
    """<table><thead><tr><th>A</th><th>B</th><th>C</th></tr></thead>
<tbody>
<tr><td>1</td><td></td><td>3</td></tr>
<tr><td></td><td>x</td><td>   </td></tr>
</tbody></table>""",
    {"logical_rows": 3, "logical_cols": 3,
     "must_contain": ["A", "B", "C", "1", "3", "x"],
     "note": "empty-cell preservation — are blanks kept as empty pipe cells so column alignment holds?"},
)

# T10 — cells containing block content (lists, <br>, links)
case(
    "t10_block_in_cell",
    "Cells containing a <ul> list, a <br> line break, and a link.",
    """<table><thead><tr><th>Feature</th><th>Notes</th></tr></thead>
<tbody>
<tr><td>Formats</td><td><ul><li>PDF</li><li>DOCX</li><li>XLSX</li></ul></td></tr>
<tr><td>Line<br>break</td><td>see <a href="https://example.com/x">docs</a></td></tr>
</tbody></table>""",
    {"logical_rows": 3, "logical_cols": 2,
     "must_contain": ["Feature", "Notes", "Formats", "PDF", "DOCX", "XLSX",
                      "break", "docs"],
     "note": "block content in cells — a newline inside a GFM cell breaks the table; does the converter inline/escape it?"},
)

# T11 — RTL (Arabic) content
case(
    "t11_rtl",
    "Right-to-left (Arabic) text in a table with dir=rtl.",
    """<table dir="rtl"><thead><tr><th>الاسم</th><th>المدينة</th></tr></thead>
<tbody>
<tr><td>علي</td><td>القاهرة</td></tr>
<tr><td>سارة</td><td>دبي</td></tr>
</tbody></table>""",
    {"logical_rows": 3, "logical_cols": 2,
     "must_contain": ["الاسم", "المدينة", "علي", "القاهرة", "سارة", "دبي"],
     "note": "RTL preservation — is the Arabic text kept intact and are cells in logical order?"},
)

# T12 — numbers with pipes and markdown-significant chars inside cells
case(
    "t12_pipe_in_cell",
    "Cells whose text contains pipe '|', asterisks, and backticks (GFM-significant).",
    """<table><thead><tr><th>Expr</th><th>Result</th></tr></thead>
<tbody>
<tr><td>a | b</td><td>*bold* text</td></tr>
<tr><td>`code`</td><td>x || y</td></tr>
</tbody></table>""",
    {"logical_rows": 3, "logical_cols": 2,
     "must_contain": ["Expr", "Result"],
     "adversarial": "literal '|' inside cells must be escaped or the table corrupts",
     "note": "does the converter escape in-cell pipes so the column count stays 2?"},
)

# T13 — caption + tfoot + multi-row thead
case(
    "t13_caption_tfoot",
    "Full semantic table: <caption>, 2-row <thead>, <tbody>, <tfoot> totals.",
    """<table>
<caption>Sales by Quarter</caption>
<thead>
  <tr><th rowspan="2">Region</th><th colspan="2">Half 1</th></tr>
  <tr><th>Q1</th><th>Q2</th></tr>
</thead>
<tbody>
  <tr><td>East</td><td>10</td><td>12</td></tr>
  <tr><td>West</td><td>8</td><td>9</td></tr>
</tbody>
<tfoot><tr><td>Total</td><td>18</td><td>21</td></tr></tfoot>
</table>""",
    {"logical_rows": 5, "logical_cols": 3,
     "must_contain": ["Sales by Quarter", "Region", "Q1", "Q2", "East", "West",
                      "Total", "18", "21"],
     "spanning": "rowspan header + colspan header + tfoot",
     "note": "does caption survive? does tfoot merge into the body? multi-row header collapse?"},
)


def build_tables_page():
    parts = ["<!doctype html><html><head><meta charset='utf-8'>",
             "<title>Complex table matrix</title></head><body>",
             "<h1>Complex table matrix</h1>"]
    for c in CASES:
        parts.append(f"<h2 id='{c['id']}'>{c['id']}: {c['desc']}</h2>")
        parts.append(c["html"])
    parts.append("</body></html>")
    html = "\n".join(parts)
    with open(os.path.join(FIX, "complex_tables.html"), "w", encoding="utf-8") as f:
        f.write(html)
    # also emit each case as a standalone page for per-case isolation
    for c in CASES:
        standalone = (f"<!doctype html><html><head><meta charset='utf-8'>"
                      f"<title>{c['id']}</title></head><body>{c['html']}</body></html>")
        with open(os.path.join(TAB, f"{c['id']}.html"), "w", encoding="utf-8") as f:
            f.write(standalone)
    manifest = [{"id": c["id"], "desc": c["desc"], "expected": c["expected"]}
                for c in CASES]
    with open(os.path.join(TAB, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump({"n_cases": len(CASES), "cases": manifest}, f, indent=2,
                  ensure_ascii=False)
    print(f"complex_tables.html + {len(CASES)} standalone cases + manifest.json written")


def build_large_xlsx():
    try:
        from openpyxl import Workbook
    except Exception as e:
        print("openpyxl missing, skipping xlsx:", e)
        return
    # large: 50k rows x 8 cols
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("data")
    ws.append([f"col{i}" for i in range(8)])
    for r in range(50000):
        ws.append([r, f"item{r}", (r * 7) % 1000, r / 3.0, f"tag{r % 20}",
                   bool(r % 2), (r * 13) % 100, f"note-{r}"])
    p = os.path.join(FIX, "docs", "large_sheet.xlsx")
    wb.save(p)
    print(f"large_sheet.xlsx written ({os.path.getsize(p)} bytes, 50000 rows x 8 cols)")

    # wide: 200 rows x 64 cols
    wb2 = Workbook(write_only=True)
    ws2 = wb2.create_sheet("wide")
    ws2.append([f"c{i}" for i in range(64)])
    for r in range(200):
        ws2.append([f"r{r}c{i}" for i in range(64)])
    p2 = os.path.join(FIX, "docs", "wide_sheet.xlsx")
    wb2.save(p2)
    print(f"wide_sheet.xlsx written ({os.path.getsize(p2)} bytes, 200 rows x 64 cols)")


if __name__ == "__main__":
    build_tables_page()
    build_large_xlsx()
    print("done")
